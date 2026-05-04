#!/usr/bin/env python3
"""
Excel Obligations Register & Notice Calendar Generator
Exports obligations registers and notice calendars to .xlsx format with formatting.

Usage:
    python3 excel_register.py --form fidic-red --type obligations --output obligations.xlsx
    python3 excel_register.py --form psscoc --type notices --output notices.xlsx
    python3 excel_register.py --form fidic-red --type both --output contract_admin.xlsx
    python3 excel_register.py --form fidic-red --type obligations --commencement 2026-05-11
"""

import argparse
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip3 install openpyxl")
    sys.exit(1)

# Import data from sibling scripts
import importlib.util
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

def get_obligations(form):
    mod = load_module("obligations_register", os.path.join(SCRIPT_DIR, "obligations_register.py"))
    if form not in mod.OBLIGATIONS:
        return None, None
    db = mod.OBLIGATIONS[form]
    return db["name"], db

def get_notices(form):
    mod = load_module("notice_calendar", os.path.join(SCRIPT_DIR, "notice_calendar.py"))
    if form not in mod.NOTICE_DATABASES:
        return None, None
    db = mod.NOTICE_DATABASES[form]
    return db["name"], db

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=11, color="2F5496")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, priority=None):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    if priority == "Critical":
        cell.fill = CRITICAL_FILL
    elif priority == "High":
        cell.fill = HIGH_FILL

def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

def create_obligations_sheet(wb, form, commencement=None):
    name, db = get_obligations(form)
    if not db:
        print(f"No obligations data for form '{form}'")
        return

    for party in ["contractor", "employer"]:
        items = db.get(party, [])
        if not items:
            continue
        
        ws = wb.create_sheet(title=f"{party.title()} Obligations")
        
        # Title
        ws.cell(row=1, column=1, value=f"{party.title()} Obligations Register — {name}")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}")
        ws.cell(row=2, column=1).font = SUBTITLE_FONT
        if commencement:
            ws.cell(row=2, column=4, value=f"Commencement: {commencement}")
            ws.cell(row=2, column=4).font = SUBTITLE_FONT
        
        # Headers
        headers = ["#", "Clause", "Obligation", "Timing", "Priority", "Category", "Status", "Notes / Action Required"]
        row = 4
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header(ws, row, len(headers))
        
        # Data
        for i, ob in enumerate(items, 1):
            row += 1
            values = [i, ob["clause"], ob["obligation"], ob["timing"], 
                     ob["priority"], ob["category"], "Pending", ""]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                style_data_cell(cell, ob["priority"])
        
        # Conditional formatting note
        row += 2
        ws.cell(row=row, column=1, value="Legend:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Red = Critical priority")
        ws.cell(row=row, column=1).fill = CRITICAL_FILL
        row += 1
        ws.cell(row=row, column=1, value="Yellow = High priority")
        ws.cell(row=row, column=1).fill = HIGH_FILL
        
        auto_width(ws)
        
        # Freeze panes
        ws.freeze_panes = "A5"

def create_notices_sheet(wb, form, commencement=None):
    name, db = get_notices(form)
    if not db:
        print(f"No notice data for form '{form}'")
        return
    
    ws = wb.create_sheet(title="Notice Calendar")
    
    # Title
    ws.cell(row=1, column=1, value=f"Notice Calendar — {name}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}")
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    if commencement:
        ws.cell(row=2, column=4, value=f"Commencement: {commencement}")
        ws.cell(row=2, column=4).font = SUBTITLE_FONT
    
    # Headers
    headers = ["#", "Clause", "Event / Obligation", "Notice Period", "Recipient", "Consequence of Non-Compliance", "Category"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, len(headers))
    
    notices = sorted(db["notices"], key=lambda x: (x["category"], x["clause"]))
    
    current_category = None
    i = 0
    for n in notices:
        row += 1
        i += 1
        
        # Category separator
        if n["category"] != current_category:
            current_category = n["category"]
        
        is_timebar = "TIME-BAR" in n["consequence"].upper() or "FINAL" in n["consequence"].upper()
        
        values = [i, n["clause"], n["event"], n["period"], n["recipient"], n["consequence"], n["category"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if is_timebar:
                cell.fill = CRITICAL_FILL
    
    auto_width(ws)
    ws.freeze_panes = "A5"

def main():
    parser = argparse.ArgumentParser(description="Excel Register Generator")
    parser.add_argument("--form", required=True, choices=["fidic-red", "fidic-yellow", "psscoc", "sia", "nec4"])
    parser.add_argument("--type", required=True, choices=["obligations", "notices", "both"])
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx file path")
    parser.add_argument("--commencement", help="Commencement date (YYYY-MM-DD)")
    args = parser.parse_args()
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    if args.type in ("obligations", "both"):
        create_obligations_sheet(wb, args.form, args.commencement)
    
    if args.type in ("notices", "both"):
        create_notices_sheet(wb, args.form, args.commencement)
    
    wb.save(args.output)
    print(f"Excel workbook saved to {args.output}")

if __name__ == "__main__":
    main()
